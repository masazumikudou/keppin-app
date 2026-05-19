import os
from config import STAFF_MASTER_PATH

_cache = None

def load_staff_by_name() -> dict:
    global _cache
    if _cache is not None:
        return _cache

    try:
        from openpyxl import load_workbook
        if not os.path.exists(STAFF_MASTER_PATH):
            return {}
        wb = load_workbook(STAFF_MASTER_PATH, read_only=True, data_only=True)
        ws = wb.active
        result = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            name      = str(row[1]).strip() if row[1] else ''
            member_id = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            if name:
                result[name] = {'name': name, 'slack_member_id': member_id}
        wb.close()
        _cache = result
        return result
    except Exception:
        return {}
